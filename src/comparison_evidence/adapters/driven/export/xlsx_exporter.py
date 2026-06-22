from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from comparison_evidence.domain.models.comparison_result import ComparisonResult


class XlsxExporter:
    artifact_name = "report.xlsx"

    def export(self, result: ComparisonResult, output_dir: str) -> str:
        path = Path(output_dir) / self.artifact_name
        path.parent.mkdir(parents=True, exist_ok=True)
        workbook = Workbook()
        summary = workbook.active
        summary.title = "Summary"
        _write_rows(summary, [
            ["Run ID", result.run_id],
            ["Created At", result.created_at.isoformat()],
            ["Before", result.manifest.before_label],
            ["After", result.manifest.after_label],
            ["Before Rows", result.summary.before_row_count],
            ["After Rows", result.summary.after_row_count],
            ["Matched Keys", result.summary.matched_key_count],
            ["Missing Before", result.summary.missing_before_count],
            ["Missing After", result.summary.missing_after_count],
            ["Changed Cells", result.summary.changed_cell_count],
            ["Compared Cells", result.summary.compared_cell_count],
            ["Cell Match Percent", result.summary.cell_match_percent],
            ["Warnings", " | ".join(result.warnings)],
        ])
        _write_rows(
            workbook.create_sheet("Schema Overlap"),
            [
                ["Section", "Columns"],
                ["Key columns", ", ".join(result.schema_overlap.key_columns)],
                ["Common columns", ", ".join(result.schema_overlap.common_columns)],
                ["Comparable columns", ", ".join(result.schema_overlap.comparable_columns)],
                ["Before-only columns", ", ".join(result.schema_overlap.before_only_columns)],
                ["After-only columns", ", ".join(result.schema_overlap.after_only_columns)],
                ["Excluded columns", ", ".join(result.schema_overlap.excluded_columns)],
            ],
        )
        _write_rows(
            workbook.create_sheet("Type Mismatches"),
            [["Column", "Before Type", "After Type"]]
            + [[m.column_name, m.before_type, m.after_type] for m in result.type_mismatches],
        )
        _write_rows(
            workbook.create_sheet("Column Stats"),
            [["Column", "Compared", "Matches", "Differences", "Match Percent"]]
            + [[s.column_name, s.compared_count, s.match_count, s.diff_count, s.match_percent] for s in result.column_stats],
        )
        _write_rows(
            workbook.create_sheet("Detailed Differences"),
            [["Key", "Column", "Before", "After"]]
            + [[_fmt(d.key), d.column_name, _fmt(d.before_value), _fmt(d.after_value)] for d in result.detailed_differences],
        )
        _write_rows(
            workbook.create_sheet("Missing Before"),
            [["Key", "Row"]] + [[_fmt(r.key), _fmt(r.row)] for r in result.missing_before],
        )
        _write_rows(
            workbook.create_sheet("Missing After"),
            [["Key", "Row"]] + [[_fmt(r.key), _fmt(r.row)] for r in result.missing_after],
        )
        _write_rows(
            workbook.create_sheet("Manifest"),
            [["Field", "Value"]] + [[key, _fmt(value)] for key, value in result.to_dict()["manifest"].items()],
        )
        workbook.save(path)
        return str(path)


def _write_rows(sheet: Worksheet, rows: list[list[Any]]) -> None:
    for row in rows:
        sheet.append(row)
    for column_cells in sheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        sheet.column_dimensions[column_cells[0].column_letter].width = min(max(max_length + 2, 12), 80)


def _fmt(value: Any) -> str:
    if isinstance(value, dict):
        return ", ".join(f"{key}={item}" for key, item in value.items())
    return "" if value is None else str(value)
