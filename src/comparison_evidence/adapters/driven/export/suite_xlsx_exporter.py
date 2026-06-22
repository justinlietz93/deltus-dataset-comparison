from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from comparison_evidence.adapters.driven.export.xlsx_exporter import XlsxExporter
from comparison_evidence.domain.models.comparison_suite_result import ComparisonSuiteResult


class XlsxSuiteExporter:
    artifact_name = "suite_report.xlsx"

    def export(self, suite_result: ComparisonSuiteResult, output_dir: str) -> str:
        path = Path(output_dir) / self.artifact_name
        path.parent.mkdir(parents=True, exist_ok=True)
        workbook = Workbook()
        summary = workbook.active
        summary.title = "Suite Summary"
        _write_rows(
            summary,
            [
                ["Suite ID", suite_result.suite_id],
                ["Suite Name", suite_result.manifest.suite_name],
                ["Created At", suite_result.created_at.isoformat()],
                ["Baseline", suite_result.manifest.baseline_label],
                ["Candidates", suite_result.summary.candidate_count],
                ["Pass", suite_result.summary.pass_count],
                ["Warn", suite_result.summary.warn_count],
                ["Fail", suite_result.summary.fail_count],
                ["Best Candidate", suite_result.summary.best_candidate_label or ""],
                ["Best Candidate Score", suite_result.summary.best_candidate_score or 0],
                ["Warnings", " | ".join(suite_result.warnings)],
            ],
        )
        _write_rows(
            workbook.create_sheet("Candidates"),
            [
                [
                    "Candidate",
                    "Status",
                    "Score",
                    "Changed Cells",
                    "Missing Before",
                    "Missing After",
                    "Type Mismatches",
                    "Duplicate Keys",
                    "Warnings",
                    "Compared Cells",
                    "Cell Match Percent",
                    "Run ID",
                ]
            ]
            + [
                [
                    item.candidate_label,
                    item.status,
                    item.difference_score,
                    item.changed_cell_count,
                    item.missing_before_count,
                    item.missing_after_count,
                    item.type_mismatch_count,
                    item.duplicate_key_count,
                    item.warning_count,
                    item.compared_cell_count,
                    item.cell_match_percent,
                    item.run_id,
                ]
                for item in suite_result.candidate_summaries
            ],
        )
        _write_rows(
            workbook.create_sheet("Manifest"),
            [["Field", "Value"]] + [[key, _fmt(value)] for key, value in suite_result.to_dict()["manifest"].items()],
        )
        workbook.save(path)
        comparison_exporter = XlsxExporter()
        for result in suite_result.results:
            comparison_exporter.export(result, path.parent / "comparisons" / _safe_segment(result.manifest.after_label))
        return str(path)


def _write_rows(sheet: Worksheet, rows: list[list[Any]]) -> None:
    for row in rows:
        sheet.append(row)
    for column_cells in sheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        sheet.column_dimensions[column_cells[0].column_letter].width = min(max(max_length + 2, 12), 80)


def _fmt(value: Any) -> str:
    if isinstance(value, list):
        return ", ".join(str(item) for item in value)
    return "" if value is None else str(value)


def _safe_segment(value: str) -> str:
    safe = "".join(character if character.isalnum() or character in "_-" else "_" for character in value)
    return safe.strip("_") or "candidate"
